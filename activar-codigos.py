import sys
import openpyxl as lexcel
import pymysql as cnn

#Variables globales
global gdb
global gcodigosProductos

def conectarBD():
    #Conectando a la BD
    print("")
    print("Conectando a la BD...")
    global gdb
    gdb = cnn.connect(host = 'localhost',
                     user = 'root',
                     password = '.4C3r04dm1n',
                     db = 'intranet',
                     charset = 'utf8mb4',
                     cursorclass = cnn.cursors.DictCursor)
    print("***Conexion Exitosa")


def cargarLibroExcel(libroName="codigos.xlsx", tabla="adm_productos"):
    #cargar el libro
    print("Cargando libro...")
    fexl = lexcel.load_workbook(libroName, data_only=True)
    print("Cargando pestañas...")
    global gdb
    global gcodigosProductos
    
    gcodigosProductos = fexl["Productos"]
    
    print("")
    print("Actualizando BD...")   
    
    try:
        with gdb.cursor() as cursor:            
            for fila in range (2, gcodigosProductos.max_row):
                sql = f"UPDATE {tabla} SET activo = 1 WHERE codigo = %s"
                cursor.execute(sql, (gcodigosProductos.cell(fila, 1).value))
                print("Codigo actualizado: " + str(gcodigosProductos.cell(fila,1).value))  
                gdb.commit()
        #print(cursor.rowcount, " Regsitros afectados")
    finally:
        gdb.close()
    
          

conectarBD()
cargarLibroExcel()
        